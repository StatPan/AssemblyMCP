"""Parser for Korean National Assembly API Excel specifications."""

from __future__ import annotations

import asyncio
import json
import logging
import tempfile
from dataclasses import asdict, dataclass
from pathlib import Path

import httpx
import openpyxl

logger = logging.getLogger(__name__)


@dataclass
class APIParameter:
    """Represents a single API parameter."""

    name: str
    type: str
    required: bool
    description: str

    def to_dict(self) -> dict:
        return asdict(self)

    @classmethod
    def from_dict(cls, data: dict) -> APIParameter:
        return cls(**data)


@dataclass
class APISpec:
    """Represents a parsed API specification."""

    service_id: str
    endpoint: str
    endpoint_url: str
    basic_params: list[APIParameter]
    request_params: list[APIParameter]

    def to_dict(self) -> dict:
        return {
            "service_id": self.service_id,
            "endpoint": self.endpoint,
            "endpoint_url": self.endpoint_url,
            "basic_params": [p.to_dict() for p in self.basic_params],
            "request_params": [p.to_dict() for p in self.request_params],
        }

    @classmethod
    def from_dict(cls, data: dict) -> APISpec:
        return cls(
            service_id=data["service_id"],
            endpoint=data["endpoint"],
            endpoint_url=data["endpoint_url"],
            basic_params=[APIParameter.from_dict(p) for p in data["basic_params"]],
            request_params=[APIParameter.from_dict(p) for p in data["request_params"]],
        )


class SpecParseError(Exception):
    """Raised when spec parsing fails."""

    pass


class SpecParser:
    """Parser for Excel API specification files."""

    # Excel files (.xlsx) are ZIP archives with this magic number
    EXCEL_MAGIC_NUMBERS = [
        b"PK\x03\x04",  # Standard ZIP file (used by .xlsx)
        b"PK\x05\x06",  # Empty ZIP archive
        b"PK\x07\x08",  # Spanned ZIP archive
    ]

    def __init__(self, cache_dir: Path | None = None):
        """
        Initialize the spec parser.

        Args:
            cache_dir: Directory to cache parsed JSON specs. If None, uses system temp dir.
        """
        self.cache_dir = cache_dir or Path(tempfile.gettempdir()) / "assembly_specs"
        self.cache_dir.mkdir(parents=True, exist_ok=True)

    def _is_valid_excel_file(self, content: bytes) -> bool:
        """
        Validate that the content is a valid Excel/ZIP file by checking magic numbers.
        """
        if len(content) < 4:
            return False
        return any(content[: len(magic)] == magic for magic in self.EXCEL_MAGIC_NUMBERS)

    def save_spec_json(self, spec: APISpec, output_dir: Path, filename: str | None = None) -> Path:
        """Save APISpec to a JSON file."""
        output_dir.mkdir(parents=True, exist_ok=True)
        name = filename or spec.service_id
        if not name.endswith(".json"):
            name += ".json"
        output_file = output_dir / name
        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(spec.to_dict(), f, ensure_ascii=False, indent=2)
        return output_file

    async def _download_excel_bytes(self, service_id: str, inf_seq: int = 2) -> bytes:
        """
        Download Excel specification file content into memory.
        """
        url = f"https://open.assembly.go.kr/portal/data/openapi/downloadOpenApiSpec.do?infId={service_id}&infSeq={inf_seq}"
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}

        try:
            async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
                response = await client.get(url, headers=headers)
                try:
                    response.raise_for_status()
                except httpx.HTTPStatusError as exc:
                    raise SpecParseError(
                        f"Failed to download spec for {service_id}: {exc.response.status_code}"
                    ) from exc

                content = response.content
                if len(content) < 100:
                    raise SpecParseError(f"Downloaded content too small: {len(content)} bytes")

                if not self._is_valid_excel_file(content):
                    raise SpecParseError(
                        f"Downloaded content for {service_id} is not a valid Excel file"
                    )

                logger.info(f"Downloaded spec for {service_id} ({len(content)} bytes)")
                return content

        except httpx.HTTPError as e:
            raise SpecParseError(f"Network error downloading spec for {service_id}: {e}") from e

    async def parse_spec(self, service_id: str, inf_seq: int = 2) -> APISpec:
        """
        Get API specification.
        1. Checks if JSON spec exists in cache.
        2. If not, downloads Excel to memory, parses it, saves JSON to cache, and returns it.
        """
        json_file = self.cache_dir / f"{service_id}.json"

        # 1. Try Cache
        if json_file.exists():
            try:
                with open(json_file, encoding="utf-8") as f:
                    data = json.load(f)
                    logger.debug(f"Loaded spec for {service_id} from cache")
                    return APISpec.from_dict(data)
            except Exception as e:
                logger.warning(f"Failed to load cached spec for {service_id}, re-downloading: {e}")

        # 2. Download and Parse (Stream Processing)
        excel_content = await self._download_excel_bytes(service_id, inf_seq)

        def _parse_sync(content: bytes):
            from io import BytesIO

            try:
                wb = openpyxl.load_workbook(BytesIO(content))
                ws = wb["Sheet1"]

                # Extract endpoint URL
                endpoint_url = self._extract_endpoint_url(ws)
                if not endpoint_url:
                    raise SpecParseError(f"Could not find endpoint URL in spec for {service_id}")

                endpoint = endpoint_url.split("/")[-1]

                # Extract parameters
                basic_params = []
                request_params = []

                in_basic_section = False
                in_request_section = False

                for row in ws.iter_rows(min_row=1, values_only=True):
                    if not row or not any(row):
                        continue

                    first_cell = str(row[0]) if row[0] else ""

                    if "기본인자" in first_cell:
                        in_basic_section = True
                        in_request_section = False
                        continue
                    elif "요청인자" in first_cell:
                        in_basic_section = False
                        in_request_section = True
                        continue
                    elif "출력값" in first_cell or "출력명" in first_cell:
                        break

                    if (in_basic_section or in_request_section) and len(row) >= 3 and row[1]:
                        type_str = str(row[1])
                        if "필수" in type_str or "선택" in type_str:
                            param = APIParameter(
                                name=str(row[0]),
                                type=type_str,
                                required="필수" in type_str,
                                description=str(row[2]) if row[2] else "",
                            )
                            if in_basic_section:
                                basic_params.append(param)
                            else:
                                request_params.append(param)

                spec = APISpec(
                    service_id=service_id,
                    endpoint=endpoint,
                    endpoint_url=endpoint_url,
                    basic_params=basic_params,
                    request_params=request_params,
                )

                # Save to JSON cache
                self.save_spec_json(spec, self.cache_dir)
                return spec

            except Exception as e:
                raise SpecParseError(f"Failed to parse spec for {service_id}: {e}") from e

        return await asyncio.to_thread(_parse_sync, excel_content)

    def clear_cache(self, service_id: str) -> None:
        """
        Remove cached JSON spec for a service.
        Useful for testing or forcing re-download.
        """
        json_file = self.cache_dir / f"{service_id}.json"
        if json_file.exists():
            json_file.unlink()
            logger.debug(f"Cleared cache for {service_id}")

    def _extract_endpoint_url(self, worksheet) -> str | None:
        """
        Extract endpoint URL from worksheet.

        Args:
            worksheet: openpyxl worksheet object

        Returns:
            Endpoint URL or None if not found
        """
        for row in worksheet.iter_rows(min_row=1, max_row=50, max_col=1):
            cell = row[0]
            if cell.value and "요청주소" in str(cell.value):
                # Next row should contain the URL
                next_row_value = worksheet.cell(cell.row + 1, 1).value
                if next_row_value and "https://" in str(next_row_value):
                    url = str(next_row_value).strip().replace("- ", "")
                    return url
        return None
